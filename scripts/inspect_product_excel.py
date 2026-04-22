from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path('/home/ubuntu/upload/0.商品編碼(1).xlsx')


def main() -> None:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheets = workbook.sheetnames
    first_sheet = workbook[sheets[0]]
    preview = []
    for index, row in enumerate(first_sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        preview.append(row[0])
        if index >= 15:
            break
    workbook.close()
    print(json.dumps({
        'path': str(EXCEL_PATH),
        'sheet_names': sheets,
        'first_sheet': sheets[0] if sheets else None,
        'first_column_preview': preview,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
