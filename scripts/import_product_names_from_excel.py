from __future__ import annotations

import json
import os
from collections import OrderedDict
from pathlib import Path
from urllib.parse import unquote, urlparse

import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook

PROJECT_ROOT = Path('/home/ubuntu/quality-dashboard-mvp')
EXCEL_PATH = Path('/home/ubuntu/upload/0.商品編碼(1).xlsx')
WORKSHEET_NAME = '商品編碼列表'
SSL_CA_PATH = '/etc/ssl/certs/ca-certificates.crt'


def normalize_label(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == 'unnamed: 0':
        return None
    return ' '.join(text.split())


def load_labels() -> list[str]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook[WORKSHEET_NAME]
    ordered: OrderedDict[str, None] = OrderedDict()
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        label = normalize_label(row[0])
        if label:
            ordered.setdefault(label, None)
    workbook.close()
    return list(ordered.keys())


def get_connection() -> pymysql.connections.Connection:
    load_dotenv(PROJECT_ROOT / '.env', override=False)
    database_url = os.getenv('DATABASE_URL', '').strip()
    if not database_url:
        raise RuntimeError('DATABASE_URL is not available')

    parsed = urlparse(database_url)
    return pymysql.connect(
        host=parsed.hostname or 'localhost',
        port=parsed.port or 3306,
        user=unquote(parsed.username or ''),
        password=unquote(parsed.password or ''),
        database=(parsed.path or '/').lstrip('/'),
        charset='utf8mb4',
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
        ssl={'ca': SSL_CA_PATH},
    )


def main() -> None:
    labels = load_labels()
    if not labels:
        raise RuntimeError('No product names were found in workbook')

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT COUNT(*) AS total FROM product_name_options')
            total_before = int(cursor.fetchone()['total'])

            cursor.execute('DELETE FROM product_name_options')

            values = []
            for index, label in enumerate(labels, start=1):
                values.append((label, 1, index * 10))

            cursor.executemany(
                'INSERT INTO product_name_options (label, active, sortOrder) VALUES (%s, %s, %s)',
                values,
            )
            connection.commit()

        summary = {
            'excel_path': str(EXCEL_PATH),
            'worksheet': WORKSHEET_NAME,
            'deleted_existing_labels': total_before,
            'excel_unique_labels': len(labels),
            'inserted_labels': len(labels),
            'first_inserted_labels': labels[:20],
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    finally:
        connection.close()


if __name__ == '__main__':
    main()
